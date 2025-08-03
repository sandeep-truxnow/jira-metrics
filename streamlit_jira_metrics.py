import streamlit as st
import pandas as pd
from jira import JIRA
from jira.exceptions import JIRAError
import re
import numpy as np
import io
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
import os
from collections import OrderedDict

# --- Configuration Constants (from your jira_cycle_lead_time.ipynb) ---
STATUS_INPUT = "Released, Closed"
CYCLE_STATUSES = ["In Progress", "In Review", "Ready for Testing", "In Testing"]
WORKFLOW_STATUSES = [
    "To Do", "In Progress", "Paused", "In Review", "Ready for Testing",
    "In Testing", "QA Complete", "In UAT", "In UAT Testing",
    "Ready for Release", "Released", "Closed"
]

SEARCH_OPTIONS_DISPLAY = OrderedDict([
    ("JIRA Story", 1),
    ("Team and Current Sprint", 2),
    ("Team and Duration", 3)
])

# Hardcoded TEAMS data from your ipynb file (name mapped to ID)
TEAMS_DATA = OrderedDict([
    ("A Team", "34e068f6-978d-4ad9-a4ef-3bf5eec72f65"),
    ("Avengers", "8d39d512-0220-4711-9ad0-f14fbf74a50e"),
    ("Jarvis", "1ec8443e-a42c-4613-bc88-513ee29203d0"),
    ("Mavrix", "1d8f251a-8fd9-4385-8f5f-6541c28bda19"),
    ("Phoenix", "ac9cc58b-b860-4c4d-8a4e-5a64f50c5122"),
    ("Quantum", "99b45e3f-49de-446c-b28d-25ef8e915ad6")
])

# Hardcoded DURATIONS data from your ipynb file
DURATIONS_DATA = OrderedDict([
    ("Year to Date", "startOfYear()"),
    ("Current Month", "startOfMonth()"),
    ("Last Month", "startOfMonth(-1)"),
    ("Last 2 Months", "startOfMonth(-2)"),
    ("Last 3 Months", "startOfMonth(-3)"),
    ("Last 6 Months", "startOfMonth(-6)"),
    ("Custom Date Range", "customDateRange()")
])

# --- Custom Field IDs (from previous context, used in extract_issue_meta) ---
CUSTOM_FIELD_ACCOUNT_ID = 'customfield_10267' # Not used in this report script's JQL
CUSTOM_FIELD_TEAM_ID = 'customfield_10001' # Used for JQL based on name, not ID here
CUSTOM_FIELD_STORY_POINTS_ID = 'customfield_10014' # Used in extract_issue_meta for Story Points column


# --- Initialize ALL Streamlit session state variables at the TOP LEVEL ---
if 'app_messages' not in st.session_state:
    st.session_state.app_messages = []
if 'message_placeholder_area' not in st.session_state:
    st.session_state.message_placeholder_area = None
if 'jira_conn_details' not in st.session_state:
    st.session_state.jira_conn_details = None
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False

if 'team_options_display' not in st.session_state:
    st.session_state.team_options_display = list(TEAMS_DATA.keys())
if 'duration_options_display' not in st.session_state:
    st.session_state.duration_options_display = list(DURATIONS_DATA.keys())
if 'story_point_options' not in st.session_state: # Default Fibonacci
    st.session_state.story_point_options = [1, 2, 3, 5, 8, 13, 21]

if 'selected_search_option_key' not in st.session_state:
    st.session_state.selected_search_option_key = None
if 'ticket_keys_input' not in st.session_state:
    st.session_state.ticket_keys_input = ""
if 'selected_team_name' not in st.session_state:
    st.session_state.selected_team_name = list(TEAMS_DATA.keys())[0] 
if 'selected_team_id' not in st.session_state:
    st.session_state.selected_team_id = TEAMS_DATA[list(TEAMS_DATA.keys())[0]]
if 'selected_duration_name' not in st.session_state:
    st.session_state.selected_duration_name = list(DURATIONS_DATA.keys())[0] 
if 'selected_duration_func' not in st.session_state:
    st.session_state.selected_duration_func = DURATIONS_DATA[list(DURATIONS_DATA.keys())[0]]
if 'selected_custom_start_date' not in st.session_state:
    st.session_state.selected_custom_start_date = None
if 'selected_custom_end_date' not in st.session_state:
    st.session_state.selected_custom_end_date = None

if 'generated_report_df_display' not in st.session_state:
    st.session_state.generated_report_df_display = None
if 'generated_report_file_buffer' not in st.session_state:
    st.session_state.generated_report_file_buffer = None
if 'generated_report_filename' not in st.session_state:
    st.session_state.generated_report_filename = None


# --- Global Message Handling ---
def _update_message_display_area():
    if st.session_state.message_placeholder_area is None:
        return
    placeholder_widget = st.session_state.message_placeholder_area
    placeholder_widget.empty()

    if st.session_state.app_messages:
        with placeholder_widget.container(border=True):
            for msg_type, msg_content in st.session_state.app_messages:
                if msg_type == "success": st.success(msg_content)
                elif msg_type == "info": st.info(msg_content)
                elif msg_type == "warning": st.warning(msg_content)
                elif msg_type == "error": st.error(msg_content)
            
            if st.button("Clear all notifications", key="clear_all_notifications_btn"):
                st.session_state.app_messages = []
                placeholder_widget.empty()
                st.rerun()
    else: placeholder_widget.empty()

def add_app_message(msg_type, msg_content):
    if 'app_messages' not in st.session_state: st.session_state.app_messages = []
    st.session_state.app_messages.append((msg_type, msg_content))
    if st.session_state.message_placeholder_area is not None:
        _update_message_display_area()

# --- Jira Connection Function ---
@st.cache_resource
def connect_to_jira_streamlit(url, username, api_token):
    try:
        jira_options = {'server': url}
        jira = JIRA(options=jira_options, basic_auth=(username, api_token))
        return jira
    except Exception as e:
        add_app_message("error", f"Error connecting to Jira: {e}")
        return None

# --- Data Fetching Functions (adapted for Streamlit caching and inputs) ---
@st.cache_data
def get_available_projects_streamlit(jira_url, jira_username, jira_api_token):
    jira_instance = connect_to_jira_streamlit(jira_url, jira_username, jira_api_token)
    if not jira_instance: return []
    try:
        projects = jira_instance.projects()
        project_list = [{'key': p.key, 'name': p.name} for p in projects]
        return project_list
    except JIRAError as e:
        add_app_message("error", f"Error fetching projects: {e}")
        return []
    except Exception as e:
        add_app_message("error", f"An unexpected error occurred while fetching projects: {e}")
        return []

@st.cache_data
def get_all_jira_users_streamlit(jira_url, jira_username, jira_api_token, filter_domain=None):
    jira_instance = connect_to_jira_streamlit(jira_url, jira_username, jira_api_token)
    if not jira_instance: return {}
    all_users = {}
    start_at = 0; max_results = 50

    while True:
        try:
            users_page = jira_instance.search_users(query='*', startAt=start_at, maxResults=max_results, includeInactive=False)
            if not users_page: break
            for user in users_page:
                if hasattr(user, 'accountId') and user.accountId:
                    email_lower = user.emailAddress.lower() if hasattr(user, 'emailAddress') else ''
                    
                    is_atlassian_user = False
                    if hasattr(user, 'accountType'):
                        if user.accountType and user.accountType.lower() == 'atlassian':
                            is_atlassian_user = True
                    else:
                        NON_HUMAN_KEYWORDS_FALLBACK = [
                            '[APP]', 'automation', 'bot', 'service', 'plugin', 'jira-system',
                            'addon', 'connect', 'integration', 'github', 'slack', 'webhook',
                            'migrator', 'system', 'importer', 'syncer'
                        ]
                        display_name_lower = user.displayName.lower() if hasattr(user, 'displayName') else ''
                        is_non_human_by_keyword = False
                        for keyword in NON_HUMAN_KEYWORDS_FALLBACK:
                            if keyword in display_name_lower or keyword in email_lower:
                                is_non_human_by_keyword = True
                                break
                        is_atlassian_user = not is_non_human_by_keyword
                    
                    is_matching_domain = True
                    if filter_domain:
                        if not email_lower or not email_lower.endswith(f"@{filter_domain.lower()}"): 
                            is_matching_domain = False

                    if is_atlassian_user and is_matching_domain:
                        all_users[user.accountId] = {
                            'displayName': user.displayName if hasattr(user, 'displayName') else user.accountId,
                            'emailAddress': user.emailAddress if hasattr(user, 'emailAddress') else 'N/A'
                        }
            start_at += max_results
            if len(users_page) < max_results: break
        except JIRAError as e:
            add_app_message("error", f"Error fetching users: {e}")
            break
        except Exception as e:
            add_app_message("error", f"An unexpected error occurred while fetching users: {e}")
            break
    
    filter_status_message = ""
    if filter_domain:
        filter_status_message = f" (filtered by domain '{filter_domain}')"
    
    add_app_message("info", f"Fetched {len(all_users)} active human Jira users{filter_status_message}.")
    return all_users

@st.cache_data
def get_custom_field_options_streamlit(jira_url, jira_username, jira_api_token, field_id, project_key, issue_type_name="Story"):
    jira_instance = connect_to_jira_streamlit(jira_url, jira_username, jira_api_token)
    if not jira_instance:
        add_app_message("error", "Jira instance not available to fetch custom field options.")
        return []

    options_list = []
    if not field_id:
        add_app_message("warning", f"Cannot fetch options: No field ID provided.")
        return []

    field_name = "Unknown Field"
    try:
        all_fields = jira_instance.fields()
        found_field_info = next((f for f in all_fields if f['id'] == field_id), None)

        if not found_field_info or 'schema' not in found_field_info:
            add_app_message("warning", f"Warning: Custom field with ID '{field_id}' not found among all Jira fields or lacks schema.")
            return []

        field_schema = found_field_info.get('schema')
        field_name = found_field_info.get('name', 'N/A')

        is_standard_select_list = field_schema.get('custom', '').startswith('com.atlassian.jira.plugin.system.customfieldtypes:select')

        if not is_standard_select_list:
            add_app_message("info", f"Note: Custom field '{field_name}' ({field_id}) is not a standard Select List type based on schema.")
            
        if is_standard_select_list:
            add_app_message("info", f"Attempting to fetch options for custom field '{field_name}' ({field_id}) using createmeta for project '{project_key}' and issue type '{issue_type_name}'...")
            try:
                createmeta = jira_instance.createmeta(projectKeys=project_key, issuetypeNames=issue_type_name, expand='projects.issuetypes.fields')
                
                if createmeta and 'projects' in createmeta and len(createmeta['projects']) > 0:
                    project_meta = createmeta['projects'][0]
                    if 'issuetypes' in project_meta and len(issue_type_meta := project_meta['issuetypes']) > 0: # Walrus operator for python >= 3.8
                        issue_type_meta = next((it for it in project_meta['issuetypes'] if it['name'] == issue_type_name), None)
                        if issue_type_meta and 'fields' in issue_type_meta:
                            if field_id in issue_type_meta['fields']:
                                field_meta = issue_type_meta['fields'][field_id]
                                if 'allowedValues' in field_meta and isinstance(field_meta['allowedValues'], list):
                                    for option_obj in field_meta['allowedValues']:
                                        option_value_str = None
                                        if isinstance(option_obj, dict) and 'value' in option_obj:
                                            option_value_str = option_obj['value']
                                        elif hasattr(option_obj, 'value'):
                                            option_value_str = option_obj.value
                                        elif hasattr(option_obj, 'name'):
                                            option_value_str = option_obj.name
                                        else:
                                            option_value_str = str(option_obj)
                                        
                                        if option_value_str and option_value_str not in options_list:
                                            options_list.append(option_value_str)
                                    if options_list:
                                        add_app_message("success", f"Successfully fetched {len(options_list)} options for '{field_name}' ({field_id}) via createmeta.")
                                        if field_id == CUSTOM_FIELD_STORY_POINTS_ID:
                                            filtered_options = []
                                            for opt in options_list:
                                                try:
                                                    num_opt = float(opt)
                                                    if num_opt >= 1 and num_opt == int(num_opt):
                                                        filtered_options.append(int(num_opt))
                                                except ValueError:
                                                    pass
                                            return sorted(list(set(filtered_options)))
                                        else:
                                            return sorted(list(set(options_list)))
                                    else:
                                        add_app_message("warning", f"Createmeta for '{field_name}' ({field_id}) found no 'allowedValues'.")
                                else:
                                    add_app_message("warning", f"Custom field '{field_name}' ({field_id}) in createmeta lacks 'allowedValues' or it's not a list.")
                            else:
                                add_app_message("warning", f"Custom field '{field_name}' ({field_id}) not found in createmeta fields for project '{project_key}' and issue type '{issue_type_name}'. Check if it's on the screen.")
                        else:
                            add_app_message("warning", f"Issue type '{issue_type_name}' fields not found in createmeta for project '{project_key}'.")
                    else:
                        add_app_message("warning", f"Issue type '{issue_type_name}' not found in createmeta for project '{project_key}'.")
                else:
                    add_app_message("warning", f"Project '{project_key}' metadata not found in createmeta.")

            except JIRAError as e:
                add_app_message("error", f"Createmeta failed for '{field_name}' ({field_id}) (Status {e.status_code}): {e.text}")
            except Exception as e:
                add_app_message("error", f"An unexpected error occurred while processing createmeta for '{field_id}': {e}")
        else:
            add_app_message("info", f"Skipping createmeta attempt for non-standard select list '{field_name}'.")

        add_app_message("info", f"Attempting to fetch existing options for '{field_name}' ({field_id}) via JQL search...")
        try:
            jql_query = f'project = "{project_key}" AND "{field_name}" is not EMPTY'
            issues = jira_instance.search_issues(jql_query, fields=field_id, maxResults=100)

            unique_options = set()
            for issue in issues:
                field_value = issue.fields.__dict__.get(field_id)
                if field_value:
                    if isinstance(field_value, dict) and 'value' in field_value:
                        unique_options.add(field_value['value'])
                    elif isinstance(field_value, list):
                        for item_obj in field_value:
                            item_value_str = None
                            if isinstance(item_obj, dict) and 'value' in item_obj:
                                item_value_str = item_obj['value']
                            elif hasattr(item_obj, 'value'):
                                item_value_str = item_obj.value
                            elif hasattr(item_obj, 'name'):
                                item_value_str = item_obj.name
                            else:
                                item_value_str = str(item_obj)
                            
                            if item_value_str:
                                unique_options.add(item_value_str)
                    elif hasattr(field_value, 'value'):
                        unique_options.add(field_value.value)
                    elif hasattr(field_value, 'name'):
                        unique_options.add(field_value.name)
                    else:
                        unique_options.add(str(field_value))

            options_list_from_jql = sorted(list(unique_options))
            
            if field_id == CUSTOM_FIELD_STORY_POINTS_ID:
                filtered_options = []
                for opt in options_list_from_jql:
                    try:
                        num_opt = float(opt)
                        if num_opt >= 1 and num_opt == int(num_opt):
                            filtered_options.append(int(num_opt))
                    except ValueError:
                        pass
                
                if filtered_options: 
                    add_app_message("success", f"Successfully collected {len(filtered_options)} filtered options for '{field_name}' ({field_id}) via JQL search.")
                    return sorted(list(set(filtered_options)))
                else: 
                    add_app_message("warning", f"JQL search found values for '{field_name}' ({field_id}), but none were valid integer Story Points (>=1).")
                    return []
            if options_list_from_jql:
                add_app_message("success", f"Successfully collected {len(options_list_from_jql)} options for '{field_name}' ({field_id}) via JQL search.")
                return options_list_from_jql

            add_app_message("warning", f"JQL search for '{field_name}' ({field_id}) found no existing issues with values for this field.")

        except JIRAError as e:
            add_app_message("error", f"JQL search for options failed (Status {e.status_code}): {e.text}")
        except Exception as e:
            add_app_message("error", f"An unexpected error occurred during JQL search for options for '{field_id}': {e}")


    except JIRAError as e:
        add_app_message("error", f"Error fetching custom field schema for '{field_id}': {e.status_code} - {e.text}")
    except Exception as e:
        add_app_message("error", f"An unexpected error occurred while determining custom field type or fetching options for '{field_id}': {e}")

    add_app_message("info", f"No predefined options found for custom field '{field_name}' ({field_id}).")
    add_app_message("info", f"Will allow free-text input for this field.")
    return []

# === GET ISSUES FROM JQL ===
def get_issues_by_jql(jql, jira_url, username, api_token):
    auth = HTTPBasicAuth(username, api_token)
    if not jql.strip():
        add_app_message("error", "JQL query cannot be empty.")
        st.stop()
    issue_keys = []
    start_at = 0
    max_results = 50
    while True:
        url = f"{jira_url}/rest/api/3/search"
        params = {
            "jql": jql,
            "fields": "key",
            "startAt": start_at,
            "maxResults": max_results
        }
        try:
            response = requests.get(url, auth=auth, params=params)
            response.raise_for_status()
            data = response.json()
            issues = data.get("issues", [])
            issue_keys.extend(issue['key'] for issue in issues)
            if len(issues) < max_results:
                break
            start_at += max_results
        except requests.exceptions.RequestException as e:
            add_app_message("error", f"Network or API error during JQL search: {e}")
            st.stop()
        except Exception as e:
            add_app_message("error", f"An unexpected error occurred during JQL search: {e}")
            st.stop()
    return issue_keys

# === FORMAT DURATION ===
# def format_duration(hours):
#     if hours is None: return "N/A"
#     if hours < 24: return f"{int(round(hours))} hrs"
#     days = int(hours // 24)
#     rem_hrs = int(round(hours % 24))
#     return f"{days} days" if rem_hrs == 0 else f"{days} days {rem_hrs} hrs"

def format_duration(hours):
    if hours is None:
        return "N/A"
    total_minutes = int(round(hours * 60))  # Convert to minutes
    days = total_minutes // (24 * 60)
    rem_minutes = total_minutes % (24 * 60)
    hrs = rem_minutes // 60
    mins = rem_minutes % 60

    parts = []
    if days > 0:
        parts.append(f"{days} days")
    if hrs > 0:
        parts.append(f"{hrs} hrs")
    if mins > 0 or not parts:
        parts.append(f"{mins} mins")

    return " ".join(parts)

# === GET ISSUE WITH CHANGELOG ===
def get_issue_changelog(issue_key, jira_url, username, api_token):
    auth = HTTPBasicAuth(username, api_token)
    url = f"{jira_url}/rest/api/3/issue/{issue_key}?expand=changelog"
    try:
        response = requests.get(url, auth=auth)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        add_app_message("error", f"Network or API error fetching changelog for {issue_key}: {e}")
        raise
    except Exception as e:
        add_app_message("error", f"An unexpected error occurred fetching changelog for {issue_key}: {e}")
        raise

def count_transitions(histories, from_status, to_status):
    count = 0
    for history in histories:
        for item in history['items']:
            if item['field'] == 'status':
                from_str = item['fromString']
                to_str = item['toString']
                
                if from_str == from_status and to_str == to_status:
                    print(f"Transition found: {history['created']} - {from_str} to {to_str}")
                    count += 1

    return count

def seconds_to_dhm(seconds):
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    return f"{days} days {hours} hrs {minutes} mins"

def get_logged_time(histories):
    logged_time_in_seconds = None
    for history in histories:
        for item in history['items']:
            if item['field'] == 'timespent':
                logged_time_in_seconds = item['to']
                break

    if logged_time_in_seconds is not None:
        logged_time_str = seconds_to_dhm(int(logged_time_in_seconds))

        return logged_time_str

# === EXCEL FORMATTER ===
def format_excel(df, output_file_label, cycle_threshold, lead_threshold):
    if cycle_threshold <= 0 or lead_threshold <= 0:
        raise ValueError("Cycle Time and Lead Time thresholds must be positive integers.")
    
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="JIRA Cycle Times")
    
    wb = load_workbook(output_buffer)
    ws = wb.active
    ws.title = "JIRA Cycle Times"
    ws.sheet_properties.tabColor = "1072BA"

    format_sheet(ws, df.columns, cycle_threshold, lead_threshold, output_file_label)
    
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# === FORMAT SHEET ===
def format_sheet(sheet, headers, cycle_threshold, lead_threshold, output_file_label):
    create_table(sheet)
    freeze_top_row(sheet)
    auto_adjust_column_width(sheet)
    align_headers(sheet)
    selected_team_name_for_sprints = output_file_label.split('_')[0].replace('-', ' ').title()
    highlight_current_sprint_multiline(sheet, headers, selected_team_name_for_sprints) 
    highlight_long_durations(sheet, cycle_threshold, lead_threshold)

# === TABLE CREATION AND FORMATTING ===
def create_table(sheet):
    table = Table(displayName="JIRAMetricsTable", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

# === SHEET FORMATTING FUNCTIONS ===
def freeze_top_row(sheet):
    sheet.freeze_panes = "B2"

def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 5
    set_border(sheet, f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    sheet.sheet_view.showGridLines = False

def align_headers(sheet):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(color="FFFFFF", bold=True)

def set_border(sheet, cell_range):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = thin_border

def highlight_long_durations(sheet, cycle_threshold, lead_threshold):
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    col_idx = map_columns(sheet)
    add_tooltip_comments(sheet, col_idx, cycle_threshold, lead_threshold)
    apply_story_points_gradient(sheet, col_idx)
    highlight_rows(sheet, col_idx, orange_fill, cycle_threshold, lead_threshold)
    add_legend(sheet, orange_fill, cycle_threshold, lead_threshold)

def map_columns(sheet):
    header = [cell.value for cell in sheet[1]]
    return {col: idx + 1 for idx, col in enumerate(header)}

def add_tooltip_comments(sheet, col_idx, cycle_threshold, lead_threshold):
    if col_idx.get("Cycle Time"):
        sheet.cell(row=1, column=col_idx["Cycle Time"]).comment = Comment(f"Orange if Cycle Time > {cycle_threshold // 24} days", "System")
    if col_idx.get("Lead Time"):
        sheet.cell(row=1, column=col_idx["Lead Time"]).comment = Comment(f"Orange if Lead Time > {lead_threshold // 24} days", "System")
    if col_idx.get("Story Points"):
        sheet.cell(row=1, column=col_idx["Story Points"]).comment = Comment("Green gradient: low → high Story Points", "System")

def apply_story_points_gradient(sheet, col_idx):
    sp_col = col_idx.get("Story Points")
    if sp_col:
        sp_letter = get_column_letter(sp_col)
        sp_range = f"{sp_letter}2:{sp_letter}{sheet.max_row}"
        blue_gradient = ColorScaleRule(
            start_type='min', start_color='E6F0FA',
            mid_type='percentile', mid_value=50, mid_color='4FC3F7',
            end_type='max', end_color='1565C0'
        )
        sheet.conditional_formatting.add(sp_range, blue_gradient)

def duration_to_hours(val):
    if not isinstance(val, str) or val.strip().upper() == "N/A":
        return None
    days = hrs = 0
    if m := re.search(r"(\d+)\s*days?", val):
        days = int(m.group(1))
    if m := re.search(r"(\d+)\s*hrs?", val):
        hrs = int(m.group(1))
    return days * 24 + hrs
   
def get_current_and_previous_sprints(team_name_for_sprint, base_sprint="2025.12", base_start_date_str="2025-06-11", sprint_length_days=14):
    base_year, base_sprint_num = map(int, base_sprint.split("."))
    base_start_date = datetime.strptime(base_start_date_str, "%Y-%m-%d").date()
    today = datetime.today().date()

    days_elapsed = (today - base_start_date).days
    if days_elapsed < 0:
        return (f"{team_name_for_sprint} {base_year}.{base_sprint_num:02d}", f"{team_name_for_sprint} {max(base_sprint_num - 1, 1):02d}")

    sprint_offset = days_elapsed // sprint_length_days
    current_sprint_num = base_sprint_num + sprint_offset
    current_year = base_year

    while current_sprint_num > 52:
        current_sprint_num -= 52
        current_year += 1
    
    previous_sprint_num = current_sprint_num - 1
    previous_sprint_year = current_year
    if previous_sprint_num <= 0:
        previous_sprint_num += 52
        previous_sprint_year -= 1


    return (
        f"{team_name_for_sprint} {current_year}.{current_sprint_num:02d}",
        f"{team_name_for_sprint} {previous_sprint_year}.{previous_sprint_num:02d}"
    )

def get_column_index_by_header(sheet, header_name):
    for cell in sheet[1]:
        if str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return -1

def highlight_current_sprint_multiline(sheet, headers, team_name_for_sprint):
    current_sprint_full, previous_sprint_full = get_current_and_previous_sprints(team_name_for_sprint, base_sprint="2025.12", base_start_date_str="2025-06-11")

    col_idx = get_column_index_by_header(sheet, "Sprints")
    if col_idx == -1:
        add_app_message("warning", "Sprints column not found for sprint highlighting.")
        return
    
    for row in sheet.iter_rows(min_row=2):
        cell = row[col_idx - 1]

        if not cell.value:
            continue

        sprint_values_in_cell = [s.strip() for s in str(cell.value).split(",")]
        modified = False
        updated_sprints = []

        for sprint_text in sprint_values_in_cell:
            sprint_clean = sprint_text.replace("🔶", "").replace("🔷", "").strip()

            if sprint_clean == current_sprint_full:
                updated_sprints.append(f"{sprint_clean} 🔶")
                modified = True
            elif sprint_clean == previous_sprint_full:
                updated_sprints.append(f"{sprint_clean} 🔷")
                modified = True
            else:
                updated_sprints.append(sprint_text)

        if modified:
            cell.value = ", ".join(updated_sprints)

def get_duration_hours_from_excel_cell(sheet, row, col):
    if col:
        val = sheet.cell(row=row, column=col).value
        return duration_to_hours(val) if isinstance(val, str) and val.strip().upper() == "N/A" else None
    return None

def highlight_cell(sheet, row, col, hours, threshold, fill):
    if col and hours is not None and hours > threshold:
        sheet.cell(row=row, column=col).fill = fill

def is_threshold_breached(hours, threshold):
    return hours is not None and hours >= threshold

def calculate_cycle_time_hours_from_excel(sheet, row, col_idx):
    total_hours = 0
    for status in CYCLE_STATUSES:
        col = col_idx.get(status)
        if not col: continue
        val = sheet.cell(row=row, column=col).value
        if isinstance(val, str) and val.strip().upper() != "N/A":
            hours_in_status = duration_to_hours(val)
            if hours_in_status is not None: total_hours += hours_in_status
    return total_hours if total_hours > 0 else None

def highlight_rows(sheet, col_idx, orange_fill, cycle_threshold, lead_threshold):
    cycle_time_header_col_idx = col_idx.get("Cycle Time")
    lead_time_header_col_idx = col_idx.get("Lead Time")

    for row in range(2, sheet.max_row + 1):
        current_cycle_hours = calculate_cycle_time_hours_from_excel(sheet, row, col_idx)
        current_lead_hours = get_duration_hours_from_excel_cell(sheet, row, lead_time_header_col_idx)

        highlight_cell(sheet, row, cycle_time_header_col_idx, current_cycle_hours, cycle_threshold, orange_fill)
        highlight_cell(sheet, row, lead_time_header_col_idx, current_lead_hours, lead_threshold, orange_fill)

        if should_apply_heatmap(current_cycle_hours, cycle_threshold, current_lead_hours, lead_threshold):
            breach_scope = determine_breach_scope(current_cycle_hours, cycle_threshold, current_lead_hours, lead_threshold)
            apply_workflow_heatmap(sheet, row, col_idx, breach_scope)

def should_apply_heatmap(cycle_hours, cycle_threshold, lead_hours, lead_threshold):
    return (is_threshold_breached(cycle_hours, cycle_threshold) or is_threshold_breached(lead_hours, lead_threshold))

def determine_breach_scope(cycle_hours, cycle_threshold, lead_hours, lead_threshold):
    lead_breach = is_threshold_breached(lead_hours, lead_threshold)
    cycle_breach = is_threshold_breached(cycle_hours, cycle_threshold)
    if lead_breach: return "lead"
    elif cycle_breach: return "cycle"
    return None

def apply_workflow_heatmap(sheet, row, col_idx, scope="lead"):
    if scope is None: return
    if scope == "cycle": workflow_subset = CYCLE_STATUSES
    elif scope == "lead": workflow_subset = [status for status in WORKFLOW_STATUSES if status not in {"Released", "Closed"}]
    else: workflow_subset = WORKFLOW_STATUSES

    row_durations = {}; values = []
    for status in workflow_subset:
        col = col_idx.get(status)
        if not col: continue
        val = sheet.cell(row=row, column=col).value
        hours = duration_to_hours(val)
        row_durations[status] = (col, hours)
        if hours is not None: values.append(hours)
    
    if not values: return

    min_val, max_val = min(values), max(values)
    delta = max_val - min_val if max_val != min_val else 1

    for status, (col, hours) in row_durations.items():
        if hours is None: continue
        intensity = (hours - min_val) / delta
        hex_color = calculate_heatmap_color(intensity)
        sheet.cell(row=row, column=col).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def calculate_heatmap_color(intensity):
    r = 255; g = int(200 - 120 * intensity); b = int(200 - 120 * intensity)
    r = max(0, min(255, r)); g = max(0, min(255, g)); b = max(0, min(255, b))
    return f"FF{r:02X}{g:02X}{b:02X}"

def add_legend(sheet, orange_fill, cycle_threshold, lead_threshold):
    legend_col = sheet.max_column + 2
    sheet.cell(row=1, column=legend_col, value="Legend").font = Font(bold=True, size=12, underline="single")
    legends = [
        (f"Cycle Time > {cycle_threshold // 24}d", orange_fill),
        (f"Lead Time > {lead_threshold // 24}d", orange_fill),
        ("Story Points: Light→Dark Blue", PatternFill(start_color="A9D0F5", end_color="1565C0", fill_type="solid")),
        ("Workflow: Light→Dark Red (per row, if breached)", PatternFill(start_color="FFCCCC", end_color="FF6666", fill_type="solid")),
    ]
    for i, (label, fill) in enumerate(legends, start=2):
        cell = sheet.cell(row=i, column=legend_col, value=label)
        cell.fill = fill
        cell.font = Font(bold=True)
    adjust_legend_column_width(sheet, legend_col)

def adjust_legend_column_width(sheet, legend_col):
    legend_letter = get_column_letter(legend_col)
    max_len = max((len(str(sheet.cell(row=row, column=legend_col).value)) for row in range(1, sheet.max_row + 1) if sheet.cell(row=row, column=legend_col).value), default=0)
    sheet.column_dimensions[legend_letter].width = max_len + 5

# === REPORT GENERATOR ===
def collect_metrics_streamlit(issue_keys, jira_url, username, api_token):
    all_metrics = []
    auth = HTTPBasicAuth(username, api_token)

    def process_issue(key):
        try:
            issue_data = get_issue_changelog(key, jira_url, username, api_token)
            issue_meta = extract_issue_meta(key, issue_data)
            metrics = calculate_state_durations(key, issue_data)
            all_metrics.append((issue_meta, metrics))
        except requests.exceptions.RequestException as req_e:
            add_app_message("error", f"Network error fetching issue {key}: {req_e}")
        except Exception as e:
            add_app_message("error", f"Error processing issue {key}: {e}")

    with ThreadPoolExecutor(max_workers=5) as executor:
        list(executor.map(process_issue, issue_keys))
    return all_metrics

def generate_report_streamlit(issue_keys, jira_url, username, api_token, cycle_threshold, lead_threshold, file_label, selected_team_name):
    add_app_message("info", f"Collecting metrics for {len(issue_keys)} issues. This may take a while...")
    all_metrics = collect_metrics_streamlit(issue_keys, jira_url, username, api_token)
    
    if not all_metrics:
        add_app_message("warning", "No metrics collected. Report will be empty.")
        return None, None, None

    headers = generate_headers()
    data = [create_row(meta, metrics, selected_team_name) for meta, metrics in all_metrics]
    
    df = pd.DataFrame(data, columns=headers)

    output_buffer = format_excel(df, file_label, cycle_threshold, lead_threshold)
    
    add_app_message("success", "Report data generated successfully!")
    return output_buffer, f"{file_label}.xlsx", df

# === CALCULATE DURATIONS ===
def calculate_durations(transitions, created_time, issue_key):
    durations = {}
    status_times = {"To Do": created_time}

    for _, to_status, timestamp in transitions:
        if to_status and to_status not in status_times:
            status_times[to_status] = timestamp

    ordered_statuses = sorted(
        [status for status in WORKFLOW_STATUSES if status in status_times],
        key=lambda s: status_times[s]
    )

    for i in range(len(ordered_statuses) - 1):
        curr = ordered_statuses[i]
        nxt = ordered_statuses[i + 1]
        start_time = status_times[curr]
        end_time = status_times[nxt]
        diff = (end_time - start_time).total_seconds() / 3600.0
        if diff >= 0:
            durations[curr] = diff
        else:
            add_app_message("warning", f"Negative duration between {curr} and {nxt} in issue {issue_key}.")

    if ordered_statuses:
        last_status = ordered_statuses[-1]
        if last_status not in durations:
            end_time = datetime.now(tz=status_times[last_status].tzinfo)
            diff = (end_time - status_times[last_status]).total_seconds() / 3600.0
            if diff >= 0:
                durations[last_status] = diff

    return durations

# === CALCULATE METRICS ===
def calculate_metrics(transitions, created_time):
    status_times = {}
    for _, to_status, timestamp in sorted(transitions, key=lambda x: x[2]):
        if to_status and to_status not in status_times:
            status_times[to_status] = timestamp

    lead_start = created_time
    lead_end = status_times.get("Released") or status_times.get("Closed")
    cycle_start = status_times.get("In Progress")
    cycle_end = status_times.get("QA Complete")

    lead_time = (lead_end - lead_start).total_seconds() / 3600.0 if lead_end else None
    cycle_time = (cycle_end - cycle_start).total_seconds() / 3600.0 if cycle_start and cycle_end else None
    return lead_time, cycle_time

# === PARSE CHANGELOG ===
def parse_changelog_from_history(changelog):
    transitions = []
    resolved_time = None

    for change in sorted(changelog, key=lambda x: x['created']):
        for item in change['items']:
            if item['field'] == 'status':
                from_status = item.get('fromString')
                to_status = item.get('toString')
                timestamp = datetime.strptime(change['created'], "%Y-%m-%dT%H:%M:%S.%f%z")
                
                transitions.append((from_status, to_status, timestamp))
                if to_status and to_status.lower() in ['qa complete', 'done', 'closed']:
                    resolved_time = timestamp
    return transitions, resolved_time

# === CALCULATE STATE DURATIONS ===
def calculate_state_durations(issue_key, issue_data):   
    changelog = issue_data['changelog']['histories']
    created_time = datetime.strptime(issue_data['fields']['created'], "%Y-%m-%dT%H:%M:%S.%f%z")
    transitions, resolved_time = parse_changelog_from_history(changelog)
    durations = calculate_durations(transitions, created_time, issue_key)
    lead_time, cycle_time = calculate_metrics(transitions, created_time)
    return {
        "lead_time_hours": lead_time,
        "cycle_time_hours": cycle_time,
        "durations_by_status_hours": dict(durations)
    }

# === EXTRACT ISSUE META ===
def extract_issue_meta(key, issue_data):
    fields = issue_data['fields']
    if not fields:
        add_app_message("error", f"No fields found for issue {key}.")
        return {}
    
    histories = issue_data['changelog']['histories']
    sprints_field = fields.get('customfield_10010')

    sprint_str = "N/A"
    if isinstance(sprints_field, list):        
        sprint_names = []
        sprints_field.sort(key=lambda x: x.get('id', 0), reverse=True)
        for sprint in sprints_field:
            if isinstance(sprint, dict) and 'name' in sprint:
                sprint_names.append(sprint['name'])
        sprint_str = ", ".join(sprint_names) if sprint_names else "N/A"

    # Fix: Ensure Story Points are explicitly converted to int if numeric, handle None/NaN
    story_points_value = fields.get(CUSTOM_FIELD_STORY_POINTS_ID, None) # Default to None if not found
    if story_points_value is None or (isinstance(story_points_value, float) and np.isnan(story_points_value)):
        story_points_value = "N/A" # Display "N/A" for missing values
    elif isinstance(story_points_value, (int, float)):
        story_points_value = int(story_points_value) # Cast to int
    elif isinstance(story_points_value, str) and story_points_value.replace('.','',1).isdigit():
        try:
            story_points_value = int(float(story_points_value)) # Handle '1.0' as string
        except ValueError:
            story_points_value = "N/A" # Fallback if string is not convertible to number
    else:
        story_points_value = str(story_points_value) # Keep as string if some other non-numeric type

    failed_qa_count = count_transitions(histories, "In Testing", "Rejected")
    if failed_qa_count is None:
        failed_qa_count = 0  # Default to 0 if no transitions found

    logged_time = get_logged_time(histories)    
    # st.info(f"Extracted issue meta for {key}: Type={fields['issuetype']['name']}, Summary={fields['summary']}, Assignee={fields['assignee']['displayName'] if fields['assignee'] else 'Unassigned'}, Status={fields['status']['name']}, Story Points={story_points_value}, Sprints={sprint_str}, Failed QA Count={failed_qa_count}, Logged Time={logged_time}")

    return {
        "Key": key,
        "Type": fields['issuetype']['name'],
        "Summary": fields['summary'],
        "Assignee": fields['assignee']['displayName'] if fields['assignee'] else "Unassigned",
        "Status": fields['status']['name'],
        "Story Points": story_points_value, # Use the potentially converted int/string value
        "Sprints": sprint_str,
        "Failed QA Count": failed_qa_count,
        "Logged Time": logged_time,
    }

# === GENERATE HEADERS ===
def generate_headers():
    return ["Key", "Type", "Summary", "Assignee", "Status", "Story Points", "Sprints", "Failed QA Count", "Logged Time", "Cycle Time", "Lead Time"] + WORKFLOW_STATUSES

# === CREATE ROW FOR EXPORT ===
def create_row(meta, metrics, selected_team_name):
    durations = metrics['durations_by_status_hours']
    row = {
        **meta,
        # "Failed QA Count": metrics['failed_qa_count'], 
        "Cycle Time": format_duration(metrics['cycle_time_hours']),
        "Lead Time": format_duration(metrics['lead_time_hours']),
    }
    
    # --- Fix: Embed diamond differentiators in Sprints column in DataFrame ---
    if "Sprints" in meta and selected_team_name:
        original_sprints_str = meta["Sprints"]
        if original_sprints_str != "N/A":
            current_sprint_full, previous_sprint_full = get_current_and_previous_sprints(selected_team_name)
            
            sprint_values_in_cell = [s.strip() for s in original_sprints_str.split(",")]
            updated_sprints = []
            for sprint_text in sprint_values_in_cell:
                sprint_clean = sprint_text.replace("🔶", "").replace("🔷", "").strip()

                if sprint_clean == current_sprint_full:
                    updated_sprints.append(f"{sprint_clean} 🔶")
                elif sprint_clean == previous_sprint_full:
                    updated_sprints.append(f"{sprint_clean} 🔷")
                else:
                    updated_sprints.append(sprint_text)

            row["Sprints"] = ", ".join(updated_sprints)
    # --- End Fix ---

    for status in WORKFLOW_STATUSES:
        duration_in_hours = durations.get(status)
        row[status] = format_duration(duration_in_hours)
    return row

# --- Pandas Styling Functions for UI Display ---
def highlight_breached_durations_ui(s, cycle_threshold_hours, lead_threshold_hours):
    cycle_time_hours = duration_to_hours(s.get("Cycle Time"))
    lead_time_hours = duration_to_hours(s.get("Lead Time"))

    styles = [''] * len(s)

    try:
        if "Cycle Time" in s.index:
            cycle_col_idx = s.index.get_loc("Cycle Time")
            if cycle_time_hours is not None and cycle_time_hours > cycle_threshold_hours:
                styles[cycle_col_idx] = 'background-color: #FFD580' # Orange
        
        if "Lead Time" in s.index:
            lead_col_idx = s.index.get_loc("Lead Time")
            if lead_time_hours is not None and lead_time_hours > lead_threshold_hours:
                styles[lead_col_idx] = 'background-color: #FFD580' # Orange
    except KeyError:
        pass

    return styles

def apply_workflow_heatmap_ui(s):
    workflow_cols_in_row = [col for col in WORKFLOW_STATUSES if col not in {"Released", "Closed"} and col in s.index]

    row_durations_numerical = []
    for col_name in workflow_cols_in_row: # Use filtered subset
        hours = duration_to_hours(s.get(col_name))
        if hours is not None:
            row_durations_numerical.append(hours)
    
    if not row_durations_numerical:
        return [''] * len(s)

    min_val, max_val = min(row_durations_numerical), max(row_durations_numerical)
    delta = max_val - min_val if max_val != min_val else 1

    styles = [''] * len(s)

    for col_idx, col_name in enumerate(s.index):
        if col_name in workflow_cols_in_row: # Apply style only to filtered subset columns
            hours = duration_to_hours(s.get(col_name))
            if hours is not None:
                intensity = (hours - min_val) / delta
                hex_color = calculate_heatmap_color(intensity)
                styles[col_idx] = f'background-color: #{hex_color[2:]}'
    return styles

def apply_story_points_gradient_ui(s, min_sp_data, max_sp_data):
    styles = [''] * len(s)

    if max_sp_data == min_sp_data:
        single_color_val = max(0, min(1, (min_sp_data - 1) / 20.0))
        single_hex = calculate_heatmap_color_blue_gradient(single_color_val)
        return [f'background-color: #{single_hex[2:]}'] * len(s)

    delta_sp = max_sp_data - min_sp_data
    if delta_sp == 0: delta_sp = 1

    for idx, val in enumerate(s):
        try:
            sp_value = float(val)
            if sp_value >= 1 and sp_value == int(sp_value):
                normalized_value = (sp_value - min_sp_data) / delta_sp
                normalized_value = max(0, min(1, normalized_value))
                
                hex_color = calculate_heatmap_color_blue_gradient(normalized_value)
                styles[idx] = f'background-color: #{hex_color[2:]}'
        except ValueError:
            pass
    return styles

def calculate_heatmap_color(intensity):
    r = 255; g = int(200 - 120 * intensity); b = int(200 - 120 * intensity)
    r = max(0, min(255, r)); g = max(0, min(255, g)); b = max(0, min(255, b))
    return f"FF{r:02X}{g:02X}{b:02X}"

def calculate_heatmap_color_blue_gradient(intensity):
    r_start, g_start, b_start = (230, 240, 250)
    r_end, g_end, b_end = (21, 101, 192)

    r = int(r_start + (r_end - r_start) * intensity)
    g = int(g_start + (g_end - g_start) * intensity)
    b = int(b_start + (b_end - b_start) * intensity)
    
    return f"FF{r:02X}{g:02X}{b:02X}"

# --- Main Streamlit App Layout ---
def main():
    st.set_page_config(layout="wide", page_title="Jira Cycle Time Reporter", page_icon=":bar_chart:")

    # --- Initialize ALL session state variables at the VERY TOP OF MAIN() ---
    if 'app_messages' not in st.session_state: st.session_state.app_messages = []
    if 'message_placeholder_area' not in st.session_state: st.session_state.message_placeholder_area = st.empty() 
    if 'jira_conn_details' not in st.session_state: st.session_state.jira_conn_details = None
    if 'data_loaded' not in st.session_state: st.session_state.data_loaded = False
    
    if 'team_options_display' not in st.session_state: st.session_state.team_options_display = list(TEAMS_DATA.keys())
    if 'duration_options_display' not in st.session_state: st.session_state.duration_options_display = list(DURATIONS_DATA.keys())
    if 'story_point_options' not in st.session_state: st.session_state.story_point_options = [1, 2, 3, 5, 8, 13, 21]

    if 'selected_search_option_key' not in st.session_state: st.session_state.selected_search_option_key = None
    if 'ticket_keys_input' not in st.session_state: st.session_state.ticket_keys_input = ""
    if 'selected_team_name' not in st.session_state: st.session_state.selected_team_name = None
    if 'selected_team_id' not in st.session_state: st.session_state.selected_team_id = None
    if 'selected_duration_name' not in st.session_state: st.session_state.selected_duration_name = None
    if 'selected_duration_func' not in st.session_state: st.session_state.selected_duration_func = None
    if 'selected_custom_start_date' not in st.session_state: st.session_state.selected_custom_start_date = None
    if 'selected_custom_end_date' not in st.session_state: st.session_state.selected_custom_end_date = None

    if 'generated_report_df_display' not in st.session_state: st.session_state.generated_report_df_display = None
    if 'generated_report_file_buffer' not in st.session_state: st.session_state.generated_report_file_buffer = None
    if 'generated_report_filename' not in st.session_state: st.session_state.generated_report_filename = None


    # Display global messages at the very top
    _update_message_display_area()

    st.title("Jira Cycle Time Reporter")
    st.markdown("""
    Generate detailed cycle and lead time reports from your Jira data.
    """)

    # --- Sidebar for Jira Credentials and General Report Options ---
    with st.sidebar:
        st.header("Jira Connection")
        jira_url = st.text_input("Jira URL", value=st.session_state.jira_conn_details[0] if st.session_state.jira_conn_details else "https://truxinc.atlassian.net", help="e.g., https://truxinc.atlassian.net", key="sidebar_jira_url")
        jira_username = st.text_input("Jira Email/Username", value=st.session_state.jira_conn_details[1] if st.session_state.jira_conn_details else "", help="Your Jira email or username for API access.", key="sidebar_jira_username")
        jira_api_token = st.text_input("Jira API Token", type="password", value=st.session_state.jira_conn_details[2] if st.session_state.jira_conn_details else "", help="Generate from your Atlassian account security settings.", key="sidebar_jira_api_token")
        
        st.header("Report Thresholds")
        cycle_time_threshold_days = st.number_input("Cycle Time Threshold (days)", min_value=1, value=7, step=1, key="cycle_threshold_days_input")
        lead_time_threshold_days = st.number_input("Lead Time Threshold (days)", min_value=1, value=21, step=1, key="lead_threshold_days_input")
        cycle_threshold_hours = cycle_time_threshold_days * 24
        lead_threshold_hours = lead_time_threshold_days * 24

        st.header("Connect & Verify Credentials")
        if st.button("Connect to Jira and Verify"):
            st.session_state.app_messages = [] 
            if not jira_url or not jira_username or not jira_api_token:
                add_app_message("error", "Please enter Jira URL, Username, and API Token.")
                return

            with st.spinner("Connecting to Jira..."):
                jira_instance = connect_to_jira_streamlit(jira_url, jira_username, jira_api_token)
                if jira_instance:
                    st.session_state.jira_conn_details = (jira_url, jira_username, jira_api_token)
                    st.session_state.data_loaded = True 
                    add_app_message("success", "Successfully connected to Jira!")
                else:
                    st.session_state.jira_conn_details = None
                    st.session_state.data_loaded = False
                    add_app_message("error", "Failed to connect to Jira. Please check your credentials.")
                st.rerun()

    # --- Main Content Area for Report Options ---
    if st.session_state.data_loaded:
        st.subheader("Report Parameters")

        # Create 3 columns for parameter selection
        col_search_pref, col_team_select, col_duration_select = st.columns(3)

        with col_search_pref:
            selected_search_option_name = st.selectbox(
                "Select Search Preference",
                options=list(SEARCH_OPTIONS_DISPLAY.keys()),
                index=list(SEARCH_OPTIONS_DISPLAY.values()).index(st.session_state.selected_search_option_key) if st.session_state.selected_search_option_key in SEARCH_OPTIONS_DISPLAY.values() else 0,
                key="search_option_selector",
                help="Choose how to select issues for the report."
            )
            # st.session_state.selected_search_option_key = SEARCH_OPTIONS_DISPLAY[selected_search_option_name]


        # Conditional inputs based on selected_search_option_key
        # if st.session_state.selected_search_option_key == 1: # JIRA Story
        #     with col_team_select: # Team Selection (always in the second column for this option)
        #         st.session_state.ticket_keys_input = st.text_input("Enter JIRA ticket key(s) (e.g., MAIN-22, MAIN-24)", value=st.session_state.ticket_keys_input, key="ticket_keys_widget")

            # Use a callback to handle changes and rerun if needed, or simply update state
            if st.session_state.selected_search_option_key != SEARCH_OPTIONS_DISPLAY[selected_search_option_name]:
                st.session_state.selected_search_option_key = SEARCH_OPTIONS_DISPLAY[selected_search_option_name]
                st.rerun() # Rerun to properly re-render conditional inputs in other columns

        # Conditional inputs based on selected_search_option_key
        # These widgets need to be explicitly placed within the columns
        if st.session_state.selected_search_option_key == 1: # JIRA Story
            with col_team_select: # Use the second column
                st.session_state.ticket_keys_input = st.text_input(
                    "Enter JIRA ticket key(s) (e.g., MAIN-22, MAIN-24)", 
                    value=st.session_state.ticket_keys_input, 
                    key="ticket_keys_widget"
                )
        
        elif st.session_state.selected_search_option_key in [2, 3]: # Team and Current Sprint OR Team and Duration
            with col_team_select: # Team Selection (second column for these options)
                team_names_display = list(TEAMS_DATA.keys())
                current_team_name_for_selector = st.session_state.selected_team_name
                current_team_idx = team_names_display.index(current_team_name_for_selector) if current_team_name_for_selector in team_names_display else 0
                
                # Callback to update session state (rerun will happen automatically from on_change)
                def on_team_selector_change_callback():
                    st.session_state.selected_team_name = st.session_state.team_selector_widget_key
                    st.session_state.selected_team_id = TEAMS_DATA.get(st.session_state.selected_team_name)
                    # No explicit st.rerun() here, as the widget's on_change already triggers one.

                st.selectbox(
                    "Select Team",
                    options=team_names_display,
                    index=current_team_idx,
                    key="team_selector_widget_key",
                    on_change=on_team_selector_change_callback,
                    help="Select the team to filter issues."
                )
            
            with col_duration_select: # Duration Selection (third column for option 3)
                if st.session_state.selected_search_option_key == 3:
                    duration_names = list(DURATIONS_DATA.keys())
                    current_duration_name_for_selector = st.session_state.selected_duration_name
                    current_duration_idx = duration_names.index(current_duration_name_for_selector) if current_duration_name_for_selector in duration_names else 0

                    def on_duration_selector_change_callback():
                        st.session_state.selected_duration_name = st.session_state.duration_selector_widget_key
                        st.session_state.selected_duration_func = DURATIONS_DATA.get(st.session_state.selected_duration_name)
                        # No explicit st.rerun() here

                    st.selectbox(
                        "Select Duration",
                        options=duration_names,
                        index=current_duration_idx,
                        key="duration_selector_widget_key",
                        on_change=on_duration_selector_change_callback,
                        help="Select the time duration for filtering issues."
                    )
                    
                    if st.session_state.selected_duration_name == "Custom Date Range":
                        # These date inputs appear dynamically in the third column
                        start_default = st.session_state.selected_custom_start_date if st.session_state.selected_custom_start_date else date(2023, 1, 1)
                        end_default = st.session_state.selected_custom_end_date if st.session_state.selected_custom_end_date else date.today()
                        
                        st.session_state.selected_custom_start_date = st.date_input("Start Date", value=start_default, key="start_date_input")
                        st.session_state.selected_custom_end_date = st.date_input("End Date", value=end_default, key="end_date_input")

        st.subheader("Generate Report")
        col1, col2 = st.columns([2, 2])
        with col1:
            st.markdown("Click the button below to generate the report based on your selections.")

            if st.button("Generate Report"):
                st.session_state.app_messages = [] 
                JQL_QUERY = ""
                file_label = ""
                selected_team_name_for_report = st.session_state.selected_team_name
                selected_team_id_for_report = st.session_state.selected_team_id 

                (jira_url_conn, jira_username_conn, jira_api_token_conn) = st.session_state.jira_conn_details

                if st.session_state.selected_search_option_key == 1: # JIRA Story
                    ticket_keys = st.session_state.ticket_keys_input.strip()
                    if not ticket_keys:
                        add_app_message("error", "JIRA ticket key cannot be empty for 'JIRA Story' search.")
                        st.stop()
                    JQL_QUERY = f"KEY IN ({ticket_keys})"
                    file_label = f"tickets_{ticket_keys.replace(',', '_')}_asof_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                else: # Team and Current Sprint OR Team and Duration
                    if not selected_team_name_for_report or not selected_team_id_for_report:
                        add_app_message("error", "Please select a Team.")
                        st.stop()
                    
                    if st.session_state.selected_search_option_key == 2: # Team and Current Sprint
                        JQL_QUERY = f"'Team[Team]' = \"{selected_team_id_for_report}\" AND sprint in openSprints() AND issuetype NOT IN (Sub-task) ORDER BY KEY"
                        file_label = f"{selected_team_name_for_report.lower().replace(' ', '_')}_current_sprint_asof_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    elif st.session_state.selected_search_option_key == 3: # Team and Duration
                        duration_func_str = st.session_state.selected_duration_func

                        if duration_func_str == "customDateRange()":
                            start_date = st.session_state.selected_custom_start_date
                            end_date = st.session_state.selected_custom_end_date
                            if not start_date or not end_date:
                                add_app_message("error", "Please select both Start Date and End Date for custom range.")
                                st.stop()
                            if start_date > end_date:
                                add_app_message("error", "Start Date cannot be after End Date.")
                                st.stop()
                            
                            start_date_str = start_date.strftime("%Y-%m-%d")
                            end_date_str = end_date.strftime("%Y-%m-%d")
                            JQL_QUERY = (
                                f"'Team[Team]' = \"{selected_team_id_for_report}\" AND issuetype NOT IN (Epic, Sub-task) "
                                f"AND created >= \"{start_date_str}\" AND created <= \"{end_date_str}\" "
                                f"AND status IN ({STATUS_INPUT}) ORDER BY KEY"
                            )
                            file_label = f"{selected_team_name_for_report.lower().replace(' ', '_')}_custom_{start_date_str}_to_{end_date_str}_asof_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        else:
                            JQL_QUERY = (
                                f"'Team[Team]' = \"{selected_team_id_for_report}\" AND issuetype NOT IN (Epic, Sub-task) "
                                f"AND created > {duration_func_str} AND status IN ({STATUS_INPUT}) ORDER BY KEY"
                            )   
                            duration_name_from_display = st.session_state.selected_duration_name
                            file_label = f"{selected_team_name_for_report.lower().replace(' ', '_')}_{duration_name_from_display.lower().replace(' ', '_')}_asof_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                if not JQL_QUERY:
                    add_app_message("error", "Failed to generate JQL query. Please check your selections.")
                    st.stop()

                add_app_message("info", f"Generated JQL Query: `{JQL_QUERY}`")

                with st.spinner("Fetching issues and generating report..."):
                    auth_url, auth_username, auth_api_token = st.session_state.jira_conn_details

                    issue_keys = get_issues_by_jql(JQL_QUERY, auth_url, auth_username, auth_api_token)
                    
                    if not issue_keys:
                        add_app_message("warning", "No issues found matching the JQL query. Report will be empty.")
                        st.session_state.generated_report_df_display = None
                        st.session_state.generated_report_file_buffer = None
                        st.session_state.generated_report_filename = None
                    else:
                        add_app_message("info", f"Found {len(issue_keys)} issues matching the JQL query.")
                        output_buffer, output_filename, report_df_for_display = generate_report_streamlit(
                            issue_keys, 
                            auth_url, 
                            auth_username, 
                            auth_api_token,
                            cycle_threshold_hours, 
                            lead_threshold_hours, 
                            file_label,
                            selected_team_name_for_report
                        )
                        st.session_state.generated_report_file_buffer = output_buffer
                        st.session_state.generated_report_filename = output_filename
                        st.session_state.generated_report_df_display = report_df_for_display
                        add_app_message("success", f"Report generated! Ready for download: {output_filename}")
        
        with col2:
            if st.session_state.generated_report_file_buffer:
                st.markdown("Once the report is generated, you can download it using the button below.")
                st.download_button(
                    label="Download Report Excel",
                    data=st.session_state.generated_report_file_buffer,
                    file_name=st.session_state.generated_report_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    

    # --- Display the generated report preview if available ---
    # if st.session_state.data_loaded and st.session_state.generated_report_df_display is not None:
    #     st.markdown("---")  # Add a horizontal line for separation

        if st.session_state.generated_report_df_display is not None:
            st.subheader("Generated Report Preview")
            # get number of records in the dataframe
            if st.session_state.generated_report_df_display.empty:
                st.warning("No data available to display in the report preview.")
                return
            
            df_size = st.session_state.generated_report_df_display.shape
            if df_size[0] == 0 or df_size[1] == 0:
                st.warning("No data available to display in the report preview.")
                return
            
            st.markdown(f"**Total Records:** {df_size[0]}") # Display total number of records

            # --- Fix: Convert Story Points to string to avoid decimals and handle N/A ---
            df_for_display_final = st.session_state.generated_report_df_display.copy()
            df_for_display_final.index = df_for_display_final.index + 1

            if "Story Points" in df_for_display_final.columns:
                df_for_display_final["Story Points"] = df_for_display_final["Story Points"].apply(
                    lambda x: str(int(x)) if isinstance(x, (int, float)) and not pd.isna(x) else 'N/A' # Convert to int then str, handle NaN
                )

            styled_df = df_for_display_final.style # Start styling here

            # 1. Highlight Breached Durations (Orange)
            styled_df = styled_df.apply(
                lambda s: highlight_breached_durations_ui(s, cycle_threshold_hours, lead_threshold_hours), axis=1
            )
            
            # 2. Apply Workflow Heatmap (Red Gradient)
            workflow_cols_present = [col for col in WORKFLOW_STATUSES if col in df_for_display_final.columns]
            if workflow_cols_present:
                 styled_df = styled_df.apply(lambda row: apply_workflow_heatmap_ui(row), axis=1, subset=workflow_cols_present)

            # 3. Apply Story Points Gradient (Blue Gradient)
            if "Story Points" in df_for_display_final.columns:
                temp_sp_series = df_for_display_final["Story Points"].apply(lambda x: float(str(x)) if str(x).replace('.','',1).isdigit() else np.nan)
                
                numerical_sp_values = temp_sp_series.dropna()

                if not numerical_sp_values.empty:
                    min_sp_data = numerical_sp_values.min()
                    max_sp_data = numerical_sp_values.max()
                    
                    if max_sp_data == min_sp_data:
                         single_color_val = max(0, min(1, (min_sp_data - 1) / 20.0))
                         single_hex = calculate_heatmap_color_blue_gradient(single_color_val)
                         styled_df = styled_df.apply(
                             lambda s_col: [f'background-color: #{single_hex[2:]}'] * len(s_col), 
                             subset=["Story Points"]
                         )
                    else:
                        styled_df = styled_df.apply(
                            lambda s_col: apply_story_points_gradient_ui(s_col, min_sp_data, max_sp_data), 
                            subset=["Story Points"]
                        )
            # --- End Styling ---

            st.dataframe(styled_df, use_container_width=True, column_config={
                "Key": st.column_config.Column(
                    "Key",
                    width="small",
                    help="Jira Issue Key"
                ),
                "Type": st.column_config.Column(
                    "Type",
                    width="small",
                    help="Jira Issue Type"
                )
            })

            # --- Display the legend from the image ---
            st.markdown("##### Legend")
            st.markdown(
                """
                <style>
                .legend-item {
                    display: flex;
                    align-items: center;
                    margin-bottom: 5px;
                }
                .color-box {
                    width: 20px;
                    height: 20px;
                    border: 1px solid #ccc;
                    margin-right: 10px;
                }
                </style>
                <div class="legend-item">
                    <div class="color-box" style="background-color: #FFD580;"></div>
                    <span>Cycle Time / Lead Time > threshold</span>
                </div>
                <div class="legend-item">
                    <div class="color-box" style="background-color: #1565C0;"></div>
                    <span>Story Points: Light → Dark Blue (low → high)</span>
                </div>
                <div class="legend-item">
                    <div class="color-box" style="background-color: #FF6666;"></div>
                    <span>Workflow: Light → Dark Red (per row, if breached)</span>
                </div>
                """, unsafe_allow_html=True
            )
            # --- End Legend Display ---


# === MAIN ENTRY POINT ===
if __name__ == "__main__":
    main()